from openpyxl import load_workbook


def excel_read(filename):
    """считывает финансовые операции из XLSX-файла"""

    workbook = load_workbook(filename)
    sheet = workbook.active  # Выбираем активную таблицу

    # Получаем заголовки столбцов
    headers = [cell.value for cell in sheet[1]]

    # Читаем данные начиная со второй строки
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        data.append(row_data)

    return data

print(excel_read('../data/operations.xlsx'))