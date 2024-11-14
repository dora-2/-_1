from openpyxl import load_workbook
import json


def excel_read(filename):
    """считывает финансовые операции из XLSX-файла"""

    workbook = load_workbook(filename)
    sheet = workbook.active  # Выбираем активную таблицу

    # Получаем заголовки столбцов
    headers = [cell.value for cell in sheet[1]]

    # Читаем данные начиная со второй строки
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        data.append(row_data)

    return data

def sjnklvg():
    # w = []
    c = 0
    q = 0
    l = 0
    p = 0
    s = excel_read('../data/operations.xlsx')
    for i in s:
        if i['Категория'] == 'Переводы':
            c += - i['Сумма операции']
        elif i['Категория'] == 'Наличные':
            q += - i['Сумма операции']
        elif i['Категория'] != 'Пополнения':
            l += - i['Сумма операции']
        elif i['Категория'] == 'Пополнения':
            p += i['Сумма операции']
    return f' Переводы: {c}, Наличные: {q}, Общая сумма расходов: {l}, Общая сумма поступлений.: {p}'

print(sjnklvg())
# print(excel_read('../data/operations.xlsx'))